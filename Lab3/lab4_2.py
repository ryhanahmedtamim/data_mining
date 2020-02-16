import tkinter
from pathlib import Path
import openpyxl as xl
from openpyxl import Workbook
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
from tkinter import filedialog
from operator import itemgetter
from PIL import ImageTk, Image
from numpy import percentile
import math
from skimage.feature import greycomatrix, greycoprops
import skimage
from scipy import stats
import cv2
import pandas as pd


def rgb2gray(rgb):
    gray = np.dot(rgb[...,:3],
                  [0.2989, 0.5870, 0.1140])*255
    return gray.astype(np.uint8)


def message_box(str):
   messagebox.showinfo("data mining", str)


def get_glcm(image):
    result = greycomatrix(image,[1], [0], levels=256, symmetric=True, normed=True)
    return result


def feature_extraction_sift(path, output_file_name, wb):
    all_values = []
    list_value1 = ["File Name"]

    for i in range(2, 1002):
        list_value1.append("F" + str(i - 1))
    row = 0
    sift = cv2.xfeatures2d.SIFT_create()
    for file in path.glob('*.png'):

        list_value = []
        file_name = file.name
        list_value.append(file_name)
        img = mpimg.imread(str(file))
        gray = rgb2gray(img)

        print(file_name)

        kp = sift.detect(gray, None)
        kps = sorted(kp, key=lambda x: -x.response)[:16]
        kps, dsc = sift.compute(gray, kps)

        dsc = dsc.flatten()
        needed_size = 1000
        if dsc.size < needed_size:
            dsc = np.concatenate([dsc, np.zeros(needed_size - dsc.size)])
        else:
            dsc = dsc[:needed_size]

        for d in dsc:
            list_value.append(d)
        all_values.append(list_value)
        row += 1
        print(row)

    data_frame = pd.DataFrame(all_values)
    data_frame.reindex()
    data_frame.columns = list_value1
    data_frame.to_excel(output_file_name, sheet_name="Sheet1", index=FALSE)
    message_box("Finished")


def feature_extraction_surf(path, output_file_name, wb):
    all_values = []
    list_value1 = ["File Name"]

    for i in range(1, 1001):
        list_value1.append("F" + str(i))

    row = 0
    surf = cv2.xfeatures2d.SURF_create()

    for file in path.glob('*.png'):
        list_value = []
        file_name = file.name
        list_value.append(file_name)
        print(file_name)

        img = mpimg.imread(str(file))
        gray = rgb2gray(img)

        kp = surf.detect(gray, None)
        kps = sorted(kp, key=lambda x: -x.response)[:16]
        kps, dsc = surf.compute(gray, kps)
        dsc = dsc.flatten()

        needed_size = 1000
        if dsc.size < needed_size:
            dsc = np.concatenate([dsc, np.zeros(needed_size - dsc.size)])
        else:
            dsc = dsc[:needed_size]

        for d in dsc:
            list_value.append(d)
        all_values.append(list_value)
        row += 1
        print(row)

    data_frame = pd.DataFrame(all_values)
    data_frame.reindex()
    data_frame.columns = list_value1
    data_frame.to_excel(output_file_name, sheet_name="Sheet1", index=FALSE)
    message_box("Finished")


def feature_extraction_glcm(path, output_file_name, wb):

    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell.value = "File Name"
    cell = sheet['b1']
    cell.value = "Maximum probability"
    cell = sheet['c1']
    cell.value = "Correlation"
    cell = sheet['d1']
    cell.value = "Contrast"
    cell = sheet['e1']
    cell.value = "Uniformity (Energy)"
    cell = sheet['f1']
    cell.value = "Homogeneity"
    cell = sheet['g1']
    cell.value = "Entropy"

    row = 2
    for file in path.glob('*.png'):
        file_name = file.name
        img = mpimg.imread(str(file))
        gray = rgb2gray(img)
        pixel_frequency = get_glcm(gray)

        maximum_probability = np.amax(pixel_frequency)
        correlation = greycoprops(pixel_frequency, 'correlation')[0][0]
        contrast = greycoprops(pixel_frequency, 'contrast')[0][0]
        energy = greycoprops(pixel_frequency, 'energy')[0][0]
        homogeneity = greycoprops(pixel_frequency, 'homogeneity')[0][0]
        entropy = skimage.measure.shannon_entropy(pixel_frequency)
        cell = sheet.cell(row, 1)
        cell.value = file_name
        cell = sheet.cell(row, 2)
        cell.value = maximum_probability
        cell = sheet.cell(row, 3)
        cell.value = correlation
        cell = sheet.cell(row, 4)
        cell.value = contrast
        cell = sheet.cell(row, 5)
        cell.value = energy
        cell = sheet.cell(row, 6)
        cell.value = homogeneity
        cell = sheet.cell(row, 7)
        cell.value = entropy
        row += 1
    wb.save(output_file_name)
    message_box("Finished")


def feature_extraction_mean_and_std(path, output_file_name, wb):
    # Mean, Median, Mode, Midrange

    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell.value = "File Name"
    cell = sheet['b1']
    cell.value = "Min"
    cell = sheet['c1']
    cell.value = "1st quantile"
    cell = sheet['d1']
    cell.value = "Median"
    cell = sheet['e1']
    cell.value = "3 quantile"
    cell = sheet['f1']
    cell.value = "max"
    cell = sheet['g1']
    cell.value = "std"
    # Mean, Median, Mode, Midrange
    cell = sheet['h1']
    cell.value = "Mean"
    cell = sheet['i1']
    cell.value = "Mode"
    cell = sheet['j1']
    cell.value = "Midrange"

    row = 2
    for file in path.glob('*.png'):
        file_name = file.name
        img = mpimg.imread(str(file))
        gray = rgb2gray(img)
        min_value = np.amin(gray)
        qua = percentile(gray, [25, 50, 75])
        q1 = qua[0]
        median = np.median(gray)
        q3 = qua[2]
        mx = np.amax(gray)

        std = np.std(gray)
        mn = np.mean(gray)
        mid_range = (np.amin(gray) + np.amax(gray)) / 2
        mode = stats.mode(gray.flatten())[0][0]

        cell = sheet.cell(row, 1)
        cell.value = file_name
        cell = sheet.cell(row, 2)
        cell.value = min_value
        cell = sheet.cell(row, 3)
        cell.value = q1
        cell = sheet.cell(row, 4)
        cell.value = median
        cell = sheet.cell(row, 5)
        cell.value = q3
        cell = sheet.cell(row, 6)
        cell.value = mx
        cell = sheet.cell(row, 7)
        cell.value = std
        cell = sheet.cell(row, 8)
        cell.value = mn
        cell = sheet.cell(row, 9)
        cell.value = mode
        cell = sheet.cell(row, 10)
        cell.value = mid_range

        row += 1
    wb.save(output_file_name)
    message_box("Finished")


def load_training_data():
    global training_folder_path
    filename = filedialog.askdirectory()
    training_folder_path = Path(filename)
    message_box("Training Data Loaded")


def load_test_data():
    global test_folder_path
    filename = filedialog.askdirectory()
    test_folder_path = Path(filename)
    message_box("Training Data Loaded")


def load_training_feature():
    name = askopenfilename(initialdir="",
                           filetypes=(("Text File", "*.xlsx"), ("All Files", "*.*")),
                           title="Choose a file."
                           )
    global training_feature_book
    training_feature_book = xl.load_workbook(name)
    message_box("Feature Data Loaded")


def load_test_feature():
    name = askopenfilename(initialdir="",
                           filetypes=(("Text File", "*.xlsx"), ("All Files", "*.*")),
                           title="Choose a file."
                           )
    global test_feature_book
    test_feature_book = xl.load_workbook(name)
    message_box("Feature Data Loaded")


def result_using_city_block():
    test_sheet = test_feature_book['Sheet1']
    result_workbook = xl.load_workbook('output.xlsx')
    result_sheet = result_workbook['Sheet1']
    result_cell = result_sheet.cell(1, 1)
    result_cell.value = "Test Image"
    result_cell = result_sheet.cell(1, 2)
    result_cell.value = "Output Image"

    for test_row in range(2, test_sheet.max_row + 1):

        cell = test_sheet.cell(test_row, 1)
        file_name = cell.value
        result_cell = result_sheet.cell(test_row, 1)
        result_cell.value = file_name

        distance = 99999999999999
        similar_image = ""
        train_sheet = training_feature_book['Sheet1']
        for train_row in range(2, train_sheet.max_row + 1):
            d = 0
            cell = train_sheet.cell(train_row, 1)
            file = cell.value
            for i in range(2, train_sheet.max_col + 1):
                test_cell = test_sheet.cell(test_row, i)
                train_cell = train_sheet.cell(train_row, i)
                d += abs(test_cell.value - train_cell.value)

            if d < distance:
                distance = d
                similar_image = file
        result_cell = result_sheet.cell(test_row, 2)
        result_cell.value = similar_image
    result_workbook.save("result.xlsx")
    message_box("Finished")


def result_using_canberra():
    test_sheet = test_feature_book['Sheet1']
    result_workbook = xl.load_workbook('output.xlsx')
    result_sheet = result_workbook['Sheet1']
    result_cell = result_sheet.cell(1, 1)
    result_cell.value = "Test Image"
    result_cell = result_sheet.cell(1, 2)
    result_cell.value = "Output Image"

    for test_row in range(2, test_sheet.max_row + 1):

        cell = test_sheet.cell(test_row, 1)
        file_name = cell.value
        result_cell = result_sheet.cell(test_row, 1)
        result_cell.value = file_name

        distance = 99999999999999
        similar_image = ""
        train_sheet = training_feature_book['Sheet1']
        for train_row in range(2, train_sheet.max_row +1):
            d = 0
            cell = train_sheet.cell(train_row, 1)
            file = cell.value
            for i in range(2, train_sheet.max_column + 1):
                test_cell = test_sheet.cell(test_row, i)
                train_cell = train_sheet.cell(train_row, i)
                d += (abs(test_cell.value - train_cell.value)/(.1+abs(test_cell.value)+abs(train_cell.value)))

            if d < distance:
                distance = d
                similar_image = file
        print (similar_image, file_name)
        result_cell = result_sheet.cell(test_row, 2)
        result_cell.value = similar_image
    result_workbook.save("result.xlsx")
    message_box("Finished")


top = Tk()
top.title("Data Mining")
top.geometry("900x500")

workbook = xl.load_workbook('output.xlsx')
training_feature_book = xl.load_workbook('output.xlsx')
training_feature_book_name = StringVar()
test_feature_book = xl.load_workbook('output.xlsx')

train_data_path = StringVar()
training_folder_path = Path(train_data_path.get())

test_data_path = StringVar()
test_folder_path = Path(test_data_path.get())

test_image = np.arange(20).reshape(4,5)


frame1 = Frame(top, bg='green',relief=RAISED, borderwidth=1)
frame1.pack(side=LEFT, padx=10)


load_button = Button(frame1, text = "Select Training Folder",
                     height=2,command = load_training_data)
load_button.pack( side = TOP, padx=5, pady=5)


extract_feature_button_cd_dd = Button(frame1, text="Extract CT+DD Feature",
                                      height=2, command = lambda: feature_extraction_mean_and_std(training_folder_path, 'output_of_train_data_cd_dd.xlsx', workbook))
extract_feature_button_cd_dd.pack( side = TOP, padx=5, pady=5)

extract_feature_button_GLCM = Button(frame1, text="Extract GLCM Feature",
                                     height=2, command=lambda: feature_extraction_glcm(training_folder_path, 'output_of_train_data_glcm.xlsx', workbook))
extract_feature_button_GLCM.pack(side=TOP, padx=5, pady=5)

extract_feature_button_SIFT = Button(frame1, text="Extract SIFT Feature",
                                     height=2, command=lambda: feature_extraction_sift(training_folder_path, 'output_of_train_data_sift.xlsx', workbook))
extract_feature_button_SIFT.pack( side = TOP, padx=5, pady=5)

extract_feature_button_SURF = Button(frame1, text="Extract SURF Feature",
                                     height=2, command=lambda: feature_extraction_surf(training_folder_path, 'output_of_train_data_surf.xlsx', workbook))
extract_feature_button_SURF.pack( side = TOP, padx=5, pady=5)


# middle button

frame2 = Frame(top, bg='blue', relief=RAISED, borderwidth=1)
frame2.pack(side=LEFT, padx=5, pady=5)

load_feature_button = Button(frame2, text = "Load Training Feature Data", height=2, command=load_training_feature)
load_feature_button.pack(side=TOP, padx=5, pady=5)


load_query_button = Button(frame2, text = "Select Test Data", height=2, command=load_test_data)
load_query_button.pack(side=TOP, padx=5, pady=5)


frame3 = Frame(top, bg='green', relief=RAISED, borderwidth=1)
frame3.pack(side=LEFT, padx=10)


test_extract_feature_button_cd_dd = Button(frame3, text="Extract CT+DD Feature",
                                           height=2, command = lambda: feature_extraction_mean_and_std(test_folder_path, 'output_of_test_data_cd_dd.xlsx', workbook))
test_extract_feature_button_cd_dd.pack( side = TOP, padx=5, pady=5)

test_extract_feature_button_GLCM = Button(frame3, text="Extract GLCM Feature",
                                          height=2, command=lambda: feature_extraction_glcm(test_folder_path, 'output_of_test_data_glcm.xlsx', workbook))
test_extract_feature_button_GLCM.pack(side=TOP, padx=5, pady=5)

test_extract_feature_button_SIFT = Button(frame3, text="Extract SIFT Feature",
                                          height=2, command=lambda: feature_extraction_sift(test_folder_path, 'output_of_test_data_sift.xlsx', workbook))
test_extract_feature_button_SIFT.pack( side = TOP, padx=5, pady=5)

test_extract_feature_button_SURF = Button(frame3, text="Extract SURF Feature",
                                          height=2, command=lambda: feature_extraction_surf(test_folder_path, 'output_of_test_data_surf.xlsx', workbook))
test_extract_feature_button_SURF.pack(side = TOP, padx=5, pady=5)

load_button = Button(frame3, text="Load Test data Feature Data",
                     height=2, command=load_test_feature)
load_button.pack( side = TOP, padx=5, pady=5)


# frame 4 starts from here

frame4 = Frame(top, bg='blue', relief=RAISED, borderwidth=1)
frame4.pack(side=LEFT, padx=5, pady=5)

city_block_result_button = Button(frame4, text = "City block distance", height=2, command=result_using_city_block)
city_block_result_button.pack(side=TOP, padx=5, pady=5 )


canberra_result_button = Button(frame4, text = "Canberra distance", height=2, command=result_using_canberra)
canberra_result_button.pack(side=TOP, padx=5, pady=5)

random_forest_result_button = Button(frame4, text = "Random Forest", height=2, command=load_test_feature)
random_forest_result_button.pack(side=TOP, padx=5, pady=5)

top.mainloop()
