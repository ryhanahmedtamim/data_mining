import tkinter
from pathlib import Path
import openpyxl as xl
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
from tkinter import filedialog
from operator import itemgetter
from PIL import ImageTk, Image
from numpy import percentile
import math

# import os
# import shutil


def rgb2gray(rgb):
    return (np.dot(rgb[...,:3],
                  [0.2989, 0.5870, 0.1140]))*255


def message_box(str):
   msg = messagebox.showinfo("data mining", str)


def mean_and_std(path, output_file_name, wb):

    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell.value = "File Name"
    cell = sheet['b1']
    cell.value = "Min"
    cell = sheet['c1']
    cell.value = "1st quantile"
    cell = sheet['d1']
    cell.value = "Median"
    cell = sheet['e1']
    cell.value = "3 quantile"
    cell = sheet['f1']
    cell.value = "max"
    cell = sheet['g1']
    cell.value = "variance"

    row = 2
    for file in path.glob('*.png'):
        file_name = file.name
        #file_name1 = file_name1.split('-')[0]
        # apple10
        #file_name = ''.join([i for i in file_name1 if not i.isdigit()])
        #print("ff")
        img = mpimg.imread(str(file))
        gray = rgb2gray(img)
        mn = np.amin(gray)
        qua = percentile(gray, [25, 50, 75])
        q1 = qua[0]
        median = qua[1]
        q3 = qua[2]
        mx = np.amax(gray)
        var = np.var(gray)

        cell = sheet.cell(row, 1)
        cell.value = file_name
        cell = sheet.cell(row, 2)
        cell.value = mn
        cell = sheet.cell(row, 3)
        cell.value = q1
        cell = sheet.cell(row, 4)
        cell.value = median
        cell = sheet.cell(row, 5)
        cell.value = q3
        cell = sheet.cell(row, 6)
        cell.value = mx
        cell = sheet.cell(row, 7)
        cell.value = var
        row += 1
    wb.save(output_file_name)
    message_box("Training Finished")

def loadTrainingData():
    global training_folder_path
    filename = filedialog.askdirectory()
    folder_path = Path(filename)
    #print(filename)
    message_box("Training Data Loaded")


def load_feature():
    name = askopenfilename(initialdir="",
                           filetypes=(("Text File", "*.xlsx"), ("All Files", "*.*")),
                           title="Choose a file."
                           )
    global training_feature_book
    feature_book = xl.load_workbook(name)
    message_box("Feature Data Loaded")


def load_test_image(root):
    name = askopenfilename(initialdir="",
                           filetypes=(("Text File", "*.png"), ("All Files", "*.*")),
                           title="Choose a file."
                           )
    global test_image

    win = root
    test_image = mpimg.imread(name)
    myvar = Label(win, text="Test Image : ")
    myvar.text = "Test Image : "
    myvar.grid(row=0, column=0)
    #print(r,c)
    global im
    im = Image.open(name)
    resized = im.resize((150, 150), Image.ANTIALIAS)
    tkimage = ImageTk.PhotoImage(resized)
    myvar = Label(win, image=tkimage)
    myvar.image = tkimage
    myvar.grid(row=0, column=1)
    message_box("Test image loaded")



def test_result(feature_book , path , root):
    gray_test_image = rgb2gray(test_image)
    sheet = feature_book['Sheet1']
    mn_of_test_image = np.amin(gray_test_image)
    q1_of_test_image, median_of_test_image, q3_of_test_image = \
        percentile(gray_test_image, [25, 50, 75])
    mx_of_test_image = np.amax(gray_test_image)
    var_of_test_image = np.var(gray_test_image)
    result_list = []

    for row in range(2, 12):
        cell = sheet.cell(row, 1)
        file_name = cell.value
        cell = sheet.cell(row, 2)
        mn = cell.value
        cell = sheet.cell(row, 3)
        q1 = cell.value
        cell = sheet.cell(row, 4)
        meian = cell.value
        cell = sheet.cell(row, 5)
        q3 = cell.value
        cell = sheet.cell(row, 6)
        mx = cell.value
        cell = sheet.cell(row, 7)
        var = cell.value
        c_distance = math.sqrt((mn-mn_of_test_image)**2
                      +abs(q1-q1_of_test_image)**2
                      +abs(meian-median_of_test_image)**2
                      +abs(q3-q3_of_test_image)**2
                      +abs(mx-mx_of_test_image)**2
                      +abs(var-var_of_test_image)**2)
        item = {'FileName' : file_name, 'distance' : c_distance}
        result_list.append(item)
    #print(result_list)
    #message_box(object_type)
    result_list.sort(key=itemgetter('distance'))
    #print(result_list)
    for row in range(12, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        file_name = cell.value
        cell = sheet.cell(row, 2)
        mn = cell.value
        cell = sheet.cell(row, 3)
        q1 = cell.value
        cell = sheet.cell(row, 4)
        meian = cell.value
        cell = sheet.cell(row, 5)
        q3 = cell.value
        cell = sheet.cell(row, 6)
        mx = cell.value
        cell = sheet.cell(row, 7)
        var = cell.value
        c_distance = math.sqrt(abs(mn - mn_of_test_image)**2
                      + abs(q1 - q1_of_test_image)**2
                      + abs(meian - median_of_test_image)**2
                      + abs(q3 - q3_of_test_image)**2
                      + abs(mx - mx_of_test_image)**2
                      + abs(var - var_of_test_image)**2)
        for item in result_list:
            if item['distance'] > c_distance:
                result_list[-1]['FileName'] = file_name
                result_list[-1]['distance'] = c_distance
                break
        result_list.sort(key=itemgetter('distance'))
    print(result_list)

    file_name_list = []

    for file in result_list:
        file_name_list.append(file['FileName'])
    win = root



    myvar = Label(win, text="Similar Image : ")
    myvar.text = "Similar Image : "
    myvar.grid(row=1, column=0)
    COLUMNS = 6
    image_count = 7
    #myvar = Label(win, text = "result").grid(row=2,col=3)
    for infile in path.glob('*.png'):

        if infile.name in file_name_list:

            if image_count == 12:
                image_count += 1
            image_count += 1

            r, c = divmod(image_count - 1, COLUMNS)
            im = Image.open(infile)
            resized = im.resize((120, 120), Image.ANTIALIAS)
            tkimage = ImageTk.PhotoImage(resized)
            myvar = Label(win, image=tkimage)
            myvar.image = tkimage
            myvar.grid(row=r, column=c)

top = Tk()
top.title("Data Mining")
top.geometry("800x500")

workbook = xl.load_workbook('output.xlsx')
training_feature_book = xl.load_workbook('output.xlsx')
train_data_path = StringVar()
training_folder_path = Path(train_data_path.get())
test_image = np.arange(20).reshape(4,5)

#label = Label(top, fg="dark green")
#label.place(x = 150,y = 150)
#label.pack()
#counter_label(label)
Yy = top.winfo_width()
print(Yy)
load_button = Button(top, text = "Select Training Folder",
                     height=2,command = loadTrainingData)
load_button.place(x = 50,y = 440)

extract_feature_button = Button(top, text = "Extract Feature",
                                height=2, command = lambda: mean_and_std(training_folder_path, 'output_of_train_data.xlsx', workbook))
extract_feature_button.place(x = 221,y = 440)

load_feature_button = Button(top, text = "Load Feature", height=2, command=load_feature)
load_feature_button.place(x = 352,y = 440)

load_query_button = Button(top, text = "Select Test Image", height=2, command=lambda: load_test_image(top))
load_query_button.place(x = 470,y = 440)

result_button = Button(top, text="Result", height=2, command=lambda: test_result(training_feature_book, training_folder_path, top))
result_button.place(x=620,y=440)
top.mainloop()
