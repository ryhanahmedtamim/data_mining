import tkinter
from pathlib import Path
import openpyxl as xl
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
from tkinter import filedialog
from operator import itemgetter
from PIL import ImageTk, Image
from numpy import percentile
import math
from skimage.feature import greycomatrix, greycoprops
import skimage

# import os
# import shutil


def rgb2gray(rgb):
    gray = np.dot(rgb[...,:3],
                  [0.2989, 0.5870, 0.1140])*255
    return gray.astype(int)


def message_box(str):
   msg = messagebox.showinfo("data mining", str)


def get_frequency(image):
    result = greycomatrix(image,[1], [0], levels=256, symmetric=True, normed=True)
    return result


def feature_extraction(path, output_file_name, wb):

    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell.value = "File Name"
    cell = sheet['b1']
    cell.value = "Maximum probability"
    cell = sheet['c1']
    cell.value = "Correlation"
    cell = sheet['d1']
    cell.value = "Contrast"
    cell = sheet['e1']
    cell.value = "Uniformity (Energy)"
    cell = sheet['f1']
    cell.value = "Homogeneity"
    cell = sheet['g1']
    cell.value = "Entropy"

    row = 2
    for file in path.glob('*.png'):
        file_name = file.name
        img = mpimg.imread(str(file))
        gray = rgb2gray(img)
        pixel_frequency = get_frequency(gray)

        maximum_probability = np.amax(pixel_frequency)
        correlation = greycoprops(pixel_frequency,'correlation')[0][0]
        contrast = greycoprops(pixel_frequency,'contrast')[0][0]
        energy = greycoprops(pixel_frequency, 'energy')[0][0]
        homogeneity = greycoprops(pixel_frequency,'homogeneity')[0][0]
        entropy = skimage.measure.shannon_entropy(pixel_frequency)
        cell = sheet.cell(row, 1)
        cell.value = file_name
        cell = sheet.cell(row, 2)
        cell.value = maximum_probability
        cell = sheet.cell(row, 3)
        cell.value = correlation
        cell = sheet.cell(row, 4)
        cell.value = contrast
        cell = sheet.cell(row, 5)
        cell.value = energy
        cell = sheet.cell(row, 6)
        cell.value = homogeneity
        cell = sheet.cell(row, 7)
        cell.value = entropy
        row += 1
    wb.save(output_file_name)
    message_box("Training Finished")

def loadTrainingData():
    global training_folder_path
    filename = filedialog.askdirectory()
    folder_path = Path(filename)
    #print(filename)
    message_box("Training Data Loaded")


def load_feature():
    name = askopenfilename(initialdir="",
                           filetypes=(("Text File", "*.xlsx"), ("All Files", "*.*")),
                           title="Choose a file."
                           )
    global training_feature_book
    feature_book = xl.load_workbook(name)
    message_box("Feature Data Loaded")


def load_test_image(root):
    name = askopenfilename(initialdir="",
                           filetypes=(("Text File", "*.png"), ("All Files", "*.*")),
                           title="Choose a file."
                           )
    global test_image

    win = root
    test_image = mpimg.imread(name)
    gray = rgb2gray(test_image)
    #print(type(gray))
    gray = np.array(gray,dtype=np.uint8)

    get_frequency(gray)
    myvar = Label(win, text="Test Image : ")
    myvar.text = "Test Image : "
    myvar.grid(row=0, column=0)
    global im
    im = Image.open(name)
    resized = im.resize((150, 150), Image.ANTIALIAS)
    tkimage = ImageTk.PhotoImage(resized)
    myvar = Label(win, image=tkimage)
    myvar.image = tkimage
    myvar.grid(row=0, column=1)
    message_box("Test image loaded")



def test_result(feature_book , path , root):
    gray_test_image = rgb2gray(test_image)
    sheet = feature_book['Sheet1']
    pixel_frequency = get_frequency(gray_test_image)
    result_list = []

    t_maximum_probability = np.amax(pixel_frequency)
    t_correlation = greycoprops(pixel_frequency, 'correlation')[0][0]
    t_contrast = greycoprops(pixel_frequency, 'contrast')[0][0]
    t_energy = greycoprops(pixel_frequency, 'energy')[0][0]
    t_homogeneity = greycoprops(pixel_frequency, 'homogeneity')[0][0]
    t_entropy = skimage.measure.shannon_entropy(pixel_frequency)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        file_name = cell.value
        cell = sheet.cell(row, 2)
        maximum_probability = cell.value
        cell = sheet.cell(row, 3)
        correlation = cell.value
        cell = sheet.cell(row, 4)
        contrast = cell.value
        cell = sheet.cell(row, 5)
        energy = cell.value
        cell = sheet.cell(row, 6)
        homogeneity = cell.value
        cell = sheet.cell(row, 7)
        entropy = cell.value
        c_distance = ((abs(maximum_probability-t_maximum_probability)/(abs(maximum_probability)+abs(t_maximum_probability)) )+
                      (abs(correlation-t_correlation)/(abs(correlation)+abs(t_correlation)))+
                      (abs(contrast-t_contrast)/(abs(contrast)+abs(t_contrast) ))+
                      (abs(energy- t_energy)/(abs(energy)+abs(t_energy)))+
                      (abs(homogeneity-t_homogeneity)/(abs(homogeneity)+abs(t_homogeneity)))+
                      (abs(entropy-t_entropy)/(abs(entropy)+abs(t_entropy))))
        item = {'FileName' : file_name, 'distance' : c_distance}
        result_list.append(item)

    result_list.sort(key=itemgetter('distance'))


    print(len(result_list))

    file_name_list = []
    count = 0
    for file in result_list:
        file_name_list.append(file['FileName'])
        count+=1;
        if count == 10:
            break
    win = root



    myvar = Label(win, text="Similar Image : ")
    myvar.text = "Similar Image : "
    myvar.grid(row=1, column=0)
    COLUMNS = 6
    image_count = 7
    #myvar = Label(win, text = "result").grid(row=2,col=3)
    for infile in path.glob('*.png'):

        if infile.name in file_name_list:

            if image_count == 12:
                image_count += 1
            image_count += 1

            r, c = divmod(image_count - 1, COLUMNS)
            im = Image.open(infile)
            resized = im.resize((120, 120), Image.ANTIALIAS)
            tkimage = ImageTk.PhotoImage(resized)
            myvar = Label(win, image=tkimage)
            myvar.image = tkimage
            myvar.grid(row=r, column=c)

top = Tk()
top.title("Data Mining")
top.geometry("800x500")

workbook = xl.load_workbook('output.xlsx')
training_feature_book = xl.load_workbook('output.xlsx')
train_data_path = StringVar()
training_folder_path = Path(train_data_path.get())
test_image = np.arange(20).reshape(4,5)

#label = Label(top, fg="dark green")
#label.place(x = 150,y = 150)
#label.pack()
#counter_label(label)
Yy = top.winfo_width()
print(Yy)
load_button = Button(top, text = "Select Training Folder",
                     height=2,command = loadTrainingData)
load_button.place(x = 50,y = 440)

extract_feature_button = Button(top, text = "Extract Feature",
                                height=2, command = lambda: feature_extraction(training_folder_path, 'output_of_train_data_coil.xlsx', workbook))
extract_feature_button.place(x = 221,y = 440)

load_feature_button = Button(top, text = "Load Feature", height=2, command=load_feature)
load_feature_button.place(x = 352,y = 440)

load_query_button = Button(top, text = "Select Test Image", height=2, command=lambda: load_test_image(top))
load_query_button.place(x = 470,y = 440)

result_button = Button(top, text="Result", height=2, command=lambda: test_result(training_feature_book, training_folder_path, top))
result_button.place(x=620,y=440)
top.mainloop()
